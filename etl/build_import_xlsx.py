import csv, re, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from collections import OrderedDict

# =============================================================================
# CONFIG
# =============================================================================
SRC = '/mnt/project'
OUT = '/home/claude/CLT_Sports_Science_DB_Import.xlsx'

FD_FILES = [
    f'{SRC}/forcedeckstestexport02_12_2026.csv',
    f'{SRC}/forcedeckstestexport02_19_2026.csv',
    f'{SRC}/forcedeckstestexport02_26_2026.csv',
]
NB_FILES = [
    f'{SRC}/nordbordtestexport02_12_2026.csv',
    f'{SRC}/nordbordtestexport02_18_2026.csv',
    f'{SRC}/nordbordtestexport02_24_2026.csv',
]
GPS_FILE = f'{SRC}/CLT_Weekly_GPS_onlineV2.xlsm'
TAPGRIP_FILE = f'{SRC}/TAP_GRIP.xlsx'

# Styling
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
HEADER_FILL = PatternFill('solid', fgColor='2F5496')
DATA_FONT = Font(name='Arial', size=10)
THIN_BORDER = Border(
    bottom=Side(style='thin', color='D9D9D9')
)

# =============================================================================
# HELPERS
# =============================================================================
SKIP_NAMES = {'Niner Niner', 'Average', 'Avg', 'Total', ''}
NAME_FIXES = {
    'Chris  Rivens': 'Chris Rivens',
    'Takeo  Massey': 'Takeo Massey',
    'Jaden Barnes': 'Jayden Barnes',      # GPS uses Jaden, FD/NB uses Jayden
}

def clean_name(name):
    if not name or not isinstance(name, str):
        return None
    name = ' '.join(name.strip().split())  # collapse multiple spaces
    name = NAME_FIXES.get(name, name)
    if name in SKIP_NAMES:
        return None
    return name

def last_first_to_first_last(name):
    if not name or not isinstance(name, str) or ',' not in name:
        return name
    parts = name.split(',', 1)
    return clean_name(f'{parts[1].strip()} {parts[0].strip()}')

def parse_asym(val):
    """Parse '17.3 R' into (17.3, 'R') or (None, None)"""
    if not val or not isinstance(val, str):
        return None, None
    m = re.match(r'([\d.]+)\s*([LR])', val.strip())
    if m:
        return float(m.group(1)), m.group(2)
    return None, None

def parse_date(d):
    if isinstance(d, (datetime, date)):
        return d.strftime('%Y-%m-%d') if isinstance(d, datetime) else d.isoformat()
    if not d or not isinstance(d, str):
        return None
    for fmt in ['%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d']:
        try:
            return datetime.strptime(d.strip(), fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return d

def parse_time_12h(t):
    if not t or not isinstance(t, str):
        return None
    try:
        return datetime.strptime(t.strip(), '%I:%M %p').strftime('%H:%M:%S')
    except ValueError:
        try:
            return datetime.strptime(t.strip(), '%I:%M:%S %p').strftime('%H:%M:%S')
        except ValueError:
            return t

def safe_float(v):
    if v is None or v == '':
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None

def safe_int(v):
    if v is None or v == '':
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None

def style_sheet(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 30)

# =============================================================================
# STEP 1: BUILD PLAYER MASTER LIST
# =============================================================================
print('Building player roster...')
players = OrderedDict()  # name -> {'position': X, 'position_group': Y}

POS_TO_GROUP = {
    'DT': 'DLINE', 'DE': 'DLINE', 'DL': 'DLINE',
    'LB': 'LINEBACKERS', 'ILB': 'LINEBACKERS', 'OLB': 'LINEBACKERS',
    'CB': 'DEFENSIVE BACKS', 'S': 'DEFENSIVE BACKS', 'DB': 'DEFENSIVE BACKS',
    'WR': 'RECEIVERS', 'TE': 'TIGHT ENDS',
    'OL': 'OFFENSIVE LINE', 'OT': 'OFFENSIVE LINE', 'OG': 'OFFENSIVE LINE', 'C': 'OFFENSIVE LINE',
    'RB': 'RUNNING BACKS', 'FB': 'RUNNING BACKS',
    'QB': 'QUARTERBACKS',
    'K': 'SPECIALISTS', 'P': 'SPECIALISTS', 'LS': 'SPECIALISTS',
}

def register_player(name, pos=None):
    name = clean_name(name)
    if not name:
        return
    if name not in players:
        players[name] = {'position': None, 'position_group': None}
    if pos and isinstance(pos, str) and pos.strip():
        pos = pos.strip()
        if players[name]['position'] is None:
            players[name]['position'] = pos
            players[name]['position_group'] = POS_TO_GROUP.get(pos, None)

# ForceDecks
for f in FD_FILES:
    with open(f, encoding='utf-8-sig') as fh:
        for row in csv.DictReader(fh):
            register_player(row['Name'])

# NordBord
for f in NB_FILES:
    with open(f, encoding='utf-8-sig') as fh:
        for row in csv.DictReader(fh):
            register_player(row['Name'])

# GPS (has position)
gps_wb = openpyxl.load_workbook(GPS_FILE, data_only=True)
for sn in gps_wb.sheetnames:
    ws = gps_wb[sn]
    headers = [c.value for c in ws[1]]
    if 'Name' in headers and 'Position Name' in headers:
        ni, pi = headers.index('Name'), headers.index('Position Name')
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[ni] and isinstance(row[ni], str):
                register_player(row[ni].strip(), row[pi])

# TAP_GRIP + Bodyweights (Last, First format)
tg_wb = openpyxl.load_workbook(TAPGRIP_FILE)
for sn in ['GRIP', 'Tap']:
    ws = tg_wb[sn]
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0] and isinstance(row[0], str) and ',' in row[0]:
            name = last_first_to_first_last(row[0])
            register_player(name, row[1])

bw_ws = gps_wb['BODYWEIGHTS']
for row in bw_ws.iter_rows(min_row=2, values_only=True):
    if row[0] and isinstance(row[0], str) and ',' in row[0]:
        name = last_first_to_first_last(row[0])
        register_player(name, row[1])

# Assign IDs
player_id_map = {}
pid = 1
for name in sorted(players.keys()):
    player_id_map[name] = pid
    pid += 1

print(f'  Found {len(player_id_map)} players')

# =============================================================================
# STEP 2: CREATE WORKBOOK
# =============================================================================
wb = openpyxl.Workbook()

# --- PLAYERS SHEET ---
print('Writing players sheet...')
ws_players = wb.active
ws_players.title = 'players'
ws_players.append(['player_id', 'player_name', 'position', 'position_group'])
for name in sorted(players.keys()):
    ws_players.append([
        player_id_map[name],
        name,
        players[name]['position'],
        players[name]['position_group']
    ])
style_sheet(ws_players, 4)
auto_width(ws_players)

# --- BODYWEIGHTS SHEET ---
print('Writing bodyweights sheet...')
ws_bw = wb.create_sheet('bodyweights')
ws_bw.append(['bodyweight_id', 'player_id', 'weigh_date', 'weight_lbs'])

bw_dates = [c.value for c in bw_ws[1] if isinstance(c.value, datetime)]
bw_id = 1
for row in bw_ws.iter_rows(min_row=2, values_only=True):
    raw_name = row[0]
    if not raw_name or not isinstance(raw_name, str) or ',' not in raw_name:
        continue
    name = last_first_to_first_last(raw_name)
    if not name or name not in player_id_map:
        continue
    p_id = player_id_map[name]
    for i, dt in enumerate(bw_dates):
        weight = row[i + 2]  # offset for name + position cols
        if weight and isinstance(weight, (int, float)):
            ws_bw.append([bw_id, p_id, dt.strftime('%Y-%m-%d'), weight])
            bw_id += 1

style_sheet(ws_bw, 4)
auto_width(ws_bw)
print(f'  {bw_id - 1} bodyweight records')

# --- CMJ_TESTS SHEET ---
print('Writing cmj_tests sheet...')
ws_cmj = wb.create_sheet('cmj_tests')
CMJ_COLS = [
    'cmj_id','player_id','test_date','test_time','bodyweight_kg','reps',
    'additional_load_lbs','jump_height_in','rsi_modified',
    'concentric_peak_force_n','eccentric_peak_force_n',
    'concentric_peak_force_per_bm','vertical_velocity_takeoff',
    'ecc_braking_rfd_asym_pct','ecc_braking_rfd_asym_side',
    'ecc_braking_rfd_L','ecc_braking_rfd_R',
    'countermovement_depth_cm',
    'positive_impulse_asym_pct','positive_impulse_asym_side',
    'braking_phase_duration_ms','flight_time_contraction_time'
]
ws_cmj.append(CMJ_COLS)

cmj_id = 1
for f in FD_FILES:
    with open(f, encoding='utf-8-sig') as fh:
        for row in csv.DictReader(fh):
            name = clean_name(row['Name'])
            if not name or name not in player_id_map:
                continue
            ecc_pct, ecc_side = parse_asym(row.get('Eccentric Braking RFD % (Asym) (%)'))
            imp_pct, imp_side = parse_asym(row.get('Positive Impulse % (Asym) (%)'))
            ws_cmj.append([
                cmj_id,
                player_id_map[name],
                parse_date(row.get('Date')),
                parse_time_12h(row.get('Time')),
                safe_float(row.get('BW [KG]')),
                safe_int(row.get('Reps')),
                safe_int(row.get('Additional Load [lb]')),
                safe_float(row.get('Jump Height (Imp-Mom) in Inches [in] ')),
                safe_float(row.get('RSI-modified [m/s] ')),
                safe_int(row.get('Concentric Peak Force [N] ')),
                safe_int(row.get('Eccentric Peak Force [N] ')),
                safe_float(row.get('Concentric Peak Force / BM [N/kg] ')),
                safe_float(row.get('Vertical Velocity at Takeoff [m/s] ')),
                ecc_pct, ecc_side,
                safe_int(row.get('Eccentric Braking RFD [N/s] (L)')),
                safe_int(row.get('Eccentric Braking RFD [N/s] (R)')),
                safe_float(row.get('Countermovement Depth [cm] ')),
                imp_pct, imp_side,
                safe_int(row.get('Braking Phase Duration [ms] ')),
                safe_float(row.get('Flight Time:Contraction Time ')),
            ])
            cmj_id += 1

style_sheet(ws_cmj, len(CMJ_COLS))
auto_width(ws_cmj)
print(f'  {cmj_id - 1} CMJ records')

# --- NORDBORD_TESTS SHEET ---
print('Writing nordbord_tests sheet...')
ws_nord = wb.create_sheet('nordbord_tests')
NORD_COLS = [
    'nord_id','player_id','test_date','test_time','device','test_type',
    'reps_L','reps_R',
    'max_force_L','max_force_R','max_imbalance_pct',
    'avg_force_L','avg_force_R','avg_imbalance_pct',
    'max_torque_L','max_torque_R',
    'max_impulse_L','max_impulse_R','impulse_imbalance_pct',
    'min_time_to_peak_L','min_time_to_peak_R',
    'avg_time_to_peak_L','avg_time_to_peak_R',
    'max_force_per_kg_L','max_force_per_kg_R',
    'avg_force_per_kg_L','avg_force_per_kg_R',
    'max_torque_per_kg_L','max_torque_per_kg_R',
    'avg_torque_per_kg_L','avg_torque_per_kg_R',
]
ws_nord.append(NORD_COLS)

nord_id = 1
for f in NB_FILES:
    with open(f, encoding='utf-8-sig') as fh:
        for row in csv.DictReader(fh):
            name = clean_name(row['Name'])
            if not name or name not in player_id_map:
                continue
            ws_nord.append([
                nord_id,
                player_id_map[name],
                parse_date(row.get('Date UTC')),
                parse_time_12h(row.get('Time UTC')),
                row.get('Device','').strip(),
                row.get('Test','').strip(),
                safe_int(row.get('L Reps')),
                safe_int(row.get('R Reps')),
                safe_float(row.get('L Max Force (N)')),
                safe_float(row.get('R Max Force (N)')),
                safe_float(row.get('Max Imbalance (%)')),
                safe_float(row.get('L Avg Force (N)')),
                safe_float(row.get('R Avg Force (N)')),
                safe_float(row.get('Avg Imbalance (%)')),
                safe_float(row.get('L Max Torque (Nm)')),
                safe_float(row.get('R Max Torque (Nm)')),
                safe_float(row.get('L Max Impulse (Ns)')),
                safe_float(row.get('R Max Impulse (Ns)')),
                safe_float(row.get('Impulse Imbalance (%)')),
                safe_float(row.get('L Min Time to Peak Force (s)')),
                safe_float(row.get('R Min Time to PeakForce (s)')),
                safe_float(row.get('L Avg Time to Peak Force (s)')),
                safe_float(row.get('R Avg Time to Peak Force (s)')),
                safe_float(row.get('L Max Force Per Kg (N/kg)')),
                safe_float(row.get('R Max Force Per Kg (N/kg)')),
                safe_float(row.get('L Avg Force Per Kg (N/kg)')),
                safe_float(row.get('R Avg Force Per Kg (N/kg)')),
                safe_float(row.get('L Max Torque Per Kg (Nm/kg)')),
                safe_float(row.get('R Max Torque Per Kg (Nm/kg)')),
                safe_float(row.get('L Avg Torque Per Kg (Nm/kg)')),
                safe_float(row.get('R Avg Torque Per Kg (Nm/kg)')),
            ])
            nord_id += 1

style_sheet(ws_nord, len(NORD_COLS))
auto_width(ws_nord)
print(f'  {nord_id - 1} NordBord records')

# --- GPS_SESSIONS SHEET ---
print('Writing gps_sessions sheet...')
ws_gps = wb.create_sheet('gps_sessions')
GPS_COLS = [
    'gps_id','player_id','session_date','activity_name',
    'avg_player_load','player_load_per_min','total_distance_y','walking_pct',
    'max_velocity_mph','pct_max_velocity',
    'accel_efforts','max_acceleration','decel_efforts','max_deceleration',
    'accel_decel_efforts','explosive_efforts',
    'velocity_band_1_y','velocity_band_2_y','velocity_band_3_y',
    'velocity_band_4_y','velocity_band_5_y','velocity_band_6_y',
    'velocity_band_7_y','velocity_band_8_y',
    'velocity_band_7_plus_efforts','relative_hsd','relative_vhsd'
]
ws_gps.append(GPS_COLS)

gps_id = 1
# Individual player dump sheets (have Position Name column)
INDIV_SHEETS = ['Week6Dump', 'wk7indy']
for sn in INDIV_SHEETS:
    if sn not in gps_wb.sheetnames:
        continue
    ws = gps_wb[sn]
    headers = [c.value for c in ws[1]]
    h = {v: i for i, v in enumerate(headers) if v}
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[h.get('Name', 1)]
        if not name or not isinstance(name, str):
            continue
        name = clean_name(name)
        if not name or name not in player_id_map:
            continue
        
        dt = row[h.get('Date', 3)]
        date_str = dt.strftime('%Y-%m-%d') if isinstance(dt, datetime) else parse_date(dt)
        
        def g(col_name):
            idx = h.get(col_name)
            return row[idx] if idx is not None and idx < len(row) else None
        
        ws_gps.append([
            gps_id,
            player_id_map[name],
            date_str,
            g('Activity Name'),
            safe_float(g('Average Player Load (Session)')),
            safe_float(g('Player Load Per Minute')),
            safe_float(g('Total Distance (y)')),
            safe_float(g('Walking %')),
            safe_float(g('Maximum Velocity (mph)')),
            safe_float(g('Percentage of Max Velocity')),
            safe_int(g('Acceleration B1-3 Average Efforts (Session) (Gen 2)')),
            safe_float(g('Max Acceleration')),
            safe_int(g('Deceleration B1-3 Average Efforts (Session) (Gen 2)')),
            safe_float(g('Max Deceleration')),
            safe_int(g('Accel + Decel Efforts')),
            safe_int(g('Explosive Efforts (IMA) (avg)')),
            safe_float(g('Velocity Band 1 Total Distance (y)')),
            safe_float(g('Velocity Band 2 Total Distance (y)')),
            safe_float(g('Velocity Band 3 Total Distance (y)')),
            safe_float(g('Velocity Band 4 Total Distance (y)')),
            safe_float(g('Velocity Band 5 Total Distance (y)')),
            safe_float(g('Velocity Band 6 Total Distance (y)')),
            safe_float(g('Velocity Band 7 Total Distance (y)')),
            safe_float(g('Velocity Band 8 Total Distance (y)')),
            safe_int(g('Velocity B7+ Total # Efforts (Gen 2) (Set 2)')),
            safe_float(g('Relative HSD (>75%)')),
            safe_float(g('Relative VHSD (>90%)')),
        ])
        gps_id += 1

style_sheet(ws_gps, len(GPS_COLS))
auto_width(ws_gps)
print(f'  {gps_id - 1} GPS session records')

# --- GRIP_TESTS SHEET ---
print('Writing grip_tests sheet...')
ws_grip = wb.create_sheet('grip_tests')
ws_grip.append(['grip_id','player_id','test_date','grip_L','grip_R'])

grip_ws = tg_wb['GRIP']
# Row 2 has dates in pairs (L col, skip R col)
date_row = [c.value for c in grip_ws[2]]
grip_dates = []
i = 2
while i < len(date_row):
    if isinstance(date_row[i], datetime):
        grip_dates.append((date_row[i], i))  # (date, L_col_index)
    i += 2

grip_id = 1
for row in grip_ws.iter_rows(min_row=4, values_only=True):
    raw_name = row[0]
    if not raw_name or not isinstance(raw_name, str) or ',' not in raw_name:
        continue
    name = last_first_to_first_last(raw_name)
    if not name or name not in player_id_map:
        continue
    p_id = player_id_map[name]
    for dt, col_idx in grip_dates:
        l_val = row[col_idx] if col_idx < len(row) else None
        r_val = row[col_idx + 1] if col_idx + 1 < len(row) else None
        if (l_val and isinstance(l_val, (int, float))) or (r_val and isinstance(r_val, (int, float))):
            ws_grip.append([
                grip_id, p_id, dt.strftime('%Y-%m-%d'),
                safe_float(l_val), safe_float(r_val)
            ])
            grip_id += 1

style_sheet(ws_grip, 5)
auto_width(ws_grip)
print(f'  {grip_id - 1} grip records')

# --- TAP_TESTS SHEET ---
print('Writing tap_tests sheet...')
ws_tap = wb.create_sheet('tap_tests')
ws_tap.append(['tap_id','player_id','test_date','tap_score'])

tap_ws = tg_wb['Tap']
tap_date_row = [c.value for c in tap_ws[1]]
tap_dates = []
for i in range(2, len(tap_date_row)):
    if isinstance(tap_date_row[i], datetime):
        tap_dates.append((tap_date_row[i], i))

tap_id = 1
for row in tap_ws.iter_rows(min_row=4, values_only=True):
    raw_name = row[0]
    if not raw_name or not isinstance(raw_name, str) or ',' not in raw_name:
        continue
    name = last_first_to_first_last(raw_name)
    if not name or name not in player_id_map:
        continue
    p_id = player_id_map[name]
    for dt, col_idx in tap_dates:
        val = row[col_idx] if col_idx < len(row) else None
        if val and isinstance(val, (int, float)):
            ws_tap.append([tap_id, p_id, dt.strftime('%Y-%m-%d'), int(val)])
            tap_id += 1

style_sheet(ws_tap, 4)
auto_width(ws_tap)
print(f'  {tap_id - 1} tap records')

# =============================================================================
# SAVE
# =============================================================================
wb.save(OUT)
print(f'\nSaved to {OUT}')
